from openpyxl.styles import PatternFill
from openpyxl import load_workbook

def set_cell_color(file_path, sheet_name, cell, color):
    workbook = load_workbook(file_path)
    sheet = workbook[sheet_name]

    fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
    sheet[cell].fill = fill

    workbook.save(file_path)
    workbook.close()
